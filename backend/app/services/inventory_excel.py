from io import BytesIO
from typing import Any

from fastapi import HTTPException, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 2_000

SYSTEM_HEADERS = ("code", "name", "owner", "description", "criticality")
SERVER_HEADERS = (
    "system_code", "environment", "hostname", "ip_address", "os", "server_type",
    "role", "description", "tags", "ssh_port", "credential_name",
)


def template_response(kind: str) -> Response:
    headers = SYSTEM_HEADERS if kind == "systems" else SERVER_HEADERS
    example = (
        ("ERP", "Enterprise Resource Planning", "Business Apps", "Core ERP", "high")
        if kind == "systems"
        else ("ERP", "Production", "erp-app-01", "10.10.10.21", "Ubuntu 24.04",
              "linux", "application", "ERP application node", "erp,production", 22,
              "ERP shared SSH")
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = kind.title()
    sheet.append(headers)
    sheet.append(example)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}2"
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = max(16, len(header) + 3)
    stream = BytesIO()
    workbook.save(stream)
    return Response(
        stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="aiops-{kind}-import.xlsx"'},
    )


def read_rows(data: bytes, kind: str) -> list[tuple[int, dict[str, Any]]]:
    if len(data) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="Excel import exceeds 5 MB")
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        raw_headers = next(values)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Excel workbook could not be read") from exc
    headers = tuple(str(value or "").strip().casefold() for value in raw_headers)
    required = SYSTEM_HEADERS if kind == "systems" else SERVER_HEADERS
    missing = [header for header in required if header not in headers]
    if missing:
        raise HTTPException(
            status_code=422, detail=f"Excel template is missing columns: {', '.join(missing)}"
        )
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values_row in enumerate(values, 2):
        if row_number > MAX_IMPORT_ROWS + 1:
            raise HTTPException(status_code=422, detail="Excel import is limited to 2000 rows")
        record = {headers[index]: value for index, value in enumerate(values_row)
                  if index < len(headers)}
        if any(value not in (None, "") for value in record.values()):
            rows.append((row_number, record))
    if not rows:
        raise HTTPException(status_code=422, detail="Excel workbook contains no data rows")
    return rows
