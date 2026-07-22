from io import BytesIO

from openpyxl import load_workbook

from app.services.inventory_excel import SERVER_HEADERS, SYSTEM_HEADERS, read_rows, template_response


def test_system_and_server_templates_have_stable_import_contracts() -> None:
    for kind, expected in (("systems", SYSTEM_HEADERS), ("servers", SERVER_HEADERS)):
        response = template_response(kind)
        workbook = load_workbook(BytesIO(response.body), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows[0] == expected
        assert len(rows) == 2


def test_excel_reader_returns_numbered_records() -> None:
    response = template_response("systems")

    rows = read_rows(response.body, "systems")

    assert rows[0][0] == 2
    assert rows[0][1]["code"] == "ERP"
